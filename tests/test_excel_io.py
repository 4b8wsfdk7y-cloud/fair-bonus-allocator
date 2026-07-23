"""Tests for excel_io: template generation, parsing, and result writing."""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import Config, Department, PoolConfig
from excel_io import (
    DEPT_COLUMNS,
    POOL_FIELDS,
    generate_template,
    parse_excel_config,
    write_results_excel,
)
from sensitivity import compute_sensitivity
from tiers import calibrate_tiers
from allocator import allocate
from v2_allocator import allocate_v2, reachability_audit


def _fill_template(
    pool_overrides: dict[str, object] | None = None,
    dept_rows: list[dict[str, object]] | None = None,
) -> bytes:
    """Generate a template, fill it via openpyxl, return bytes."""
    from openpyxl import load_workbook

    data = generate_template()
    wb = load_workbook(io.BytesIO(data))

    # Fill Pool sheet
    ws_pool = wb["Pool"]
    pool_label_to_key = {label: key for key, label, _d, _n in POOL_FIELDS}
    pool_overrides = pool_overrides or {}
    for row in ws_pool.iter_rows(min_row=2):
        label = row[0].value
        if label in pool_label_to_key and pool_label_to_key[label] in pool_overrides:
            row[1].value = pool_overrides[pool_label_to_key[label]]

    # Fill Departments sheet
    ws_dept = wb["Departments"]
    dept_label_to_col = {label: idx + 1 for idx, (_k, label, _d) in enumerate(DEPT_COLUMNS)}
    dept_rows = dept_rows or []
    for i, dept in enumerate(dept_rows, start=2):
        for label, val in dept.items():
            if label in dept_label_to_col and val is not None:
                ws_dept.cell(row=i, column=dept_label_to_col[label], value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------

def test_template_has_both_sheets():
    data = generate_template()
    xls = pd.ExcelFile(io.BytesIO(data))
    assert "Pool" in xls.sheet_names
    assert "Departments" in xls.sheet_names


def test_template_pool_has_all_fields():
    data = generate_template()
    pool_df = pd.read_excel(io.BytesIO(data), sheet_name="Pool")
    labels_in_template = set(pool_df.iloc[:, 0])
    expected_labels = {label for _k, label, _d, _n in POOL_FIELDS}
    assert expected_labels.issubset(labels_in_template)


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------

def test_parse_minimal_valid_config():
    """A config with just the required fields parses successfully."""
    data = _fill_template(
        pool_overrides={
            "pool_total": 500_000,
            "profit_target": 10_000_000,
            "profit_baseline": 8_000_000,
        },
        dept_rows=[
            {"部门名称": "A", "KPI 基线": 100, "KPI Stretch": 120, "β (利润弹性)": 1000, "人头数": 10},
            {"部门名称": "B", "KPI 基线": 100, "KPI Stretch": 120, "β (利润弹性)": 1000, "人头数": 10},
        ],
    )
    cfg = parse_excel_config(data)
    assert cfg.pool.pool_total == 500_000
    assert len(cfg.departments) == 2
    assert cfg.departments[0].name == "A"
    assert cfg.departments[0].beta == 1000


def test_parse_rejects_missing_pool_sheet():
    # Build an Excel with only a Departments sheet
    buf = io.BytesIO()
    pd.DataFrame([{"name": "A"}]).to_excel(buf, sheet_name="Departments", index=False)
    with pytest.raises(ValueError, match="Pool"):
        parse_excel_config(buf.getvalue())


def test_parse_rejects_empty_departments():
    data = _fill_template(
        pool_overrides={"pool_total": 500_000, "profit_target": 10_000_000, "profit_baseline": 8_000_000},
        dept_rows=[],  # no dept rows
    )
    with pytest.raises(ValueError, match="没有数据行"):
        parse_excel_config(data)


def test_parse_rejects_missing_dept_name():
    data = _fill_template(
        pool_overrides={"pool_total": 500_000, "profit_target": 10_000_000, "profit_baseline": 8_000_000},
        dept_rows=[
            {"KPI 基线": 100, "KPI Stretch": 120, "β (利润弹性)": 1000},  # no name
        ],
    )
    with pytest.raises(ValueError, match="缺少部门名称"):
        parse_excel_config(data)


def test_parse_quota_validation_propagates():
    """Bad quota sum should raise during parse (via Config.validate_v2)."""
    data = _fill_template(
        pool_overrides={"pool_total": 500_000, "profit_target": 10_000_000, "profit_baseline": 8_000_000},
        dept_rows=[
            {"部门名称": "A", "KPI 基线": 100, "KPI Stretch": 120, "β (利润弹性)": 1000, "配额 q_d": 0.5},
            {"部门名称": "B", "KPI 基线": 100, "KPI Stretch": 120, "β (利润弹性)": 1000, "配额 q_d": 0.6},  # sum=1.1
        ],
    )
    with pytest.raises(ValueError, match="sum to 1.0"):
        parse_excel_config(data)


def test_parse_handles_optional_fields_as_none():
    """Empty cells for optional fields (CI, quota) should not crash parsing."""
    data = _fill_template(
        pool_overrides={"pool_total": 500_000, "profit_target": 10_000_000, "profit_baseline": 8_000_000},
        dept_rows=[
            {"部门名称": "A", "KPI 基线": 100, "KPI Stretch": 120, "β (利润弹性)": 1000, "人头数": 10},
            # quota / CI / ρ all blank → defaults
        ],
    )
    cfg = parse_excel_config(data)
    d = cfg.departments[0]
    assert d.quota is None
    assert d.beta_ci_lower is None
    assert d.beta_confidence_weight == 1.0  # default


def test_parse_full_example_config_roundtrip():
    """End-to-end: fill template with example_config data, parse, run v2."""
    import yaml
    with open(Path(__file__).resolve().parent.parent / "example_config.yaml") as f:
        raw = yaml.safe_load(f)

    pool_overrides = {
        "pool_total": raw["pool"]["pool_total"],
        "profit_target": raw["pool"]["profit_target"],
        "profit_baseline": raw["pool"]["profit_baseline"],
        "A 档利润份额": raw["pool"].get("theta_a", 0.15),
        "S 档利润份额": raw["pool"].get("theta_s", 0.30),
        "基础池比例 λ": raw["pool"].get("lambda_base_ratio", 0.3),
        "达成率夹断 a_max": raw["pool"].get("a_max", 1.5),
        "允许延迟池": raw["pool"].get("deferred_pool_enabled", True),
        "地板份额 (v1)": raw["pool"].get("min_pool_share", 0.0),
    }
    key_to_label = {k: l for k, l, _d in DEPT_COLUMNS}
    dept_rows = []
    for d in raw["departments"]:
        row = {}
        for key, label, _default in DEPT_COLUMNS:
            if key in d:
                row[label] = d[key]
        dept_rows.append(row)

    data = _fill_template(pool_overrides=pool_overrides, dept_rows=dept_rows)
    cfg = parse_excel_config(data)

    assert len(cfg.departments) == 8
    assert cfg.departments[0].name == "Sales"
    assert cfg.departments[0].beta == 0.08
    assert cfg.departments[0].quota == 0.20

    # v2 must run cleanly
    sens = compute_sensitivity(cfg)
    result = allocate_v2(cfg, sens)
    assert result.total_allocated + result.deferred_pool <= cfg.pool.pool_total + 1e-6


# ---------------------------------------------------------------------------
# Write results
# ---------------------------------------------------------------------------

def test_write_results_excel_has_all_sheets():
    cfg = Config(
        pool=PoolConfig(pool_total=1_000_000, profit_target=20_000_000, profit_baseline=17_000_000),
        departments=[
            Department(name="A", kpi_baseline=100, kpi_stretch=120, beta=1000,
                       headcount=10, quota=0.5),
            Department(name="B", kpi_baseline=100, kpi_stretch=120, beta=1000,
                       headcount=10, quota=0.5),
        ],
    )
    sens = compute_sensitivity(cfg)
    tiers = calibrate_tiers(cfg, sens)
    r_v1 = allocate(cfg, sens, tiers)
    r_v2 = allocate_v2(cfg, sens)
    audit = reachability_audit(cfg, sens)

    xlsx_bytes = write_results_excel(cfg, r_v2.df, audit, r_v2.release_gates, r_v1.df)
    xls = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    expected_sheets = {"Summary", "v2_Allocation", "v1_Allocation", "Reachability", "Release_Gates", "Input_Pool"}
    assert expected_sheets.issubset(set(xls.sheet_names))
